import tkinter as tk
from tkinter.filedialog import askopenfilename
from time import sleep
import openpyxl
import win32com.client
from tkinter import messagebox
from tkinter import ttk 
import base64
import os
from Icon_file import Icon_class
class filepathdata():
    file_tong_hop = ""
    file_con_path = []
class data():
    datalist = []
    #RANGE = {"row1":"","row2":"","col1":"","col2":""}
icon = Icon_class.Icon
icondata= base64.b64decode(icon)
tempFile= "icon.ico"
iconfile= open(tempFile,"wb")
iconfile.write(icondata)
iconfile.close()
    

class ImportUI:
    def __init__(self):
        self.m = tk.Tk()
        self.m.wm_iconbitmap(tempFile)
        self.m.title('Nhập file')
        button = tk.Button(self.m,text="OK",width=20,command=self.ok)
        button.pack(side="top")
        self.thf = tk.LabelFrame(self.m,text="File tổng hợp")
        self.thf.pack()
        self.fnb = tk.Button(self.thf,text="Chọn file tổng hợp",width=25,command=self.cfinalfile)   # final file button
        self.fnb.pack()
        self.fnl = tk.Label(self.thf,text="[File tổng hợp]")
        self.fnl.pack()
        self.cf = tk.LabelFrame(self.m,text="File con")
        self.cf.pack()
        self.sfb = tk.Button(self.cf,text="Chọn files con",width=25,command=self.smallfile)  # small file button
        self.sfb.pack()
        self.cl = tk.Label(self.cf,text="[File con]")
        self.cl.pack()
        
        self.running = True
        while self.running:
            self.m.update()
        self.m.destroy()

        #filename = tk.filedialog.askopenfilename() 
        #print(filename)
       
       
    def cfinalfile(self):
   
        filename = askopenfilename() 
        print(filename)
        filepathdata.file_tong_hop = filename
        self.fnl['text'] = filepathdata.file_tong_hop
    def smallfile(self):
        files = askopenfilename(multiple=True) 
        print(files)
        root = tk.Tk()
        root.withdraw()
        var = root.tk.splitlist(files)
        fileconlist = ""
        for f in var:
            filepathdata.file_con_path.append(f)
            print(f)
            fileconlist += f+"\n"
        self.cl['text'] = fileconlist
    
    def ok(self):
        if filepathdata.file_tong_hop and len(filepathdata.file_con_path) > 0:
            
            self.running = False
        else:
            messagebox.showerror("File thiếu","Vui lòng nhập đủ file")

class Main:
    def __init__(self):
        self.active = True
        self.mn = tk.Tk()
        self.mn.wm_iconbitmap(tempFile)

        self.mn.title('Phần mềm tổng hợp')
        self.filebar = tk.Menu(self.mn)
        self.tep = tk.Menu(self.filebar,tearoff=False)
        self.tep.add_command(label="Thoát",command=self.quit)
        self.filebar.add_cascade(label="Tệp",menu=self.tep)
        self.filebar.add_command(label="Chạy một lần",command=self.run_once)
        
        self.mn.config(menu=self.filebar)

        
        self.th = tk.LabelFrame(self.mn,text="Đã chọn")
        self.th.pack(side='left')
        self.buttonlist = tk.LabelFrame(self.mn,text="Hoạt động")
        self.buttonlist.pack(side='bottom')
        self.addnew = tk.Button(self.buttonlist,text="Thêm mới",width=15,command=self.addnew)
        self.addnew.pack()
        self.delete = tk.Button(self.buttonlist,text="Xóa",width=15,command=self.delete_data)
        self.delete.pack()
        self.addnew = tk.Button(self.buttonlist,text="Tổng hợp",width=15,command=self.run_all)
        self.addnew.pack()
        self.chooselist = tk.Listbox(self.th)
        self.chooselist.pack()
        #self.chooselist.bind('<Button-1>',self.changelook) 
        
        self.inf_lf = tk.LabelFrame(self.mn,text="Thông tin")
        self.info_ths = tk.Label(self.inf_lf,text="Sheet 1: ")
        self.inf_lf.pack(side='top')
        self.info_ths.pack(side='top')
        self.info_th = tk.Label(self.inf_lf,text="Vị trí 1: ")
        self.info_th.pack(side='top')
        
        self.info_cs = tk.Label(self.inf_lf,text="Sheet 2: ")
        self.info_cs.pack(side='bottom')
        self.info_c = tk.Label(self.inf_lf,text="Vị trí 2: ")
        self.info_c.pack(side='bottom')
        self.info_t = tk.Label(self.inf_lf,text="Chế độ: ")
        self.info_t.pack(side='bottom')
        wb =  openpyxl.load_workbook(filename = filepathdata.file_con_path[0])
        self.c_test_sheet = wb.active.title
        wb =  openpyxl.load_workbook(filename = filepathdata.file_tong_hop)
        self.th_test_sheet = wb.active.title


        #self.mn.mainloop()
        while self.active:
            self.changelook(self)
            self.mn.update()
    def change_pp_box(self,event):
        print(event)
        if self._count.get() == 0:
                 self.pp_count_box['state'] = 'normal'
        else:
            self.pp_count_box['state'] = 'disabled'
    def changelook(self,event):
        #print("Clicked")
     
        
        try:
            self.chooselist.get(self.chooselist.curselection())
        except:
            pass
            #print("Null")
        else:
            self.info_th['text'] = "Vị trí 1: " + data.datalist[self.chooselist.curselection()[0]][3]
            self.info_c['text'] = "Vị trí 2: " + data.datalist[self.chooselist.curselection()[0]][4]
            self.info_ths['text'] = "Sheet 1: " + data.datalist[self.chooselist.curselection()[0]][1]
            self.info_cs['text'] = "Sheet 2: " + data.datalist[self.chooselist.curselection()[0]][2]
            self.info_t['text'] = "Chế độ: " + ("Tổng" if data.datalist[self.chooselist.curselection()[0]][5] == 1 else "Đếm")
            #print(self.chooselist.curselection()[0])
    def addnew(self):
        self.newUI = tk.Tk()
        self.newUI.title('Thêm mới')
        self.newUI.wm_iconbitmap(tempFile)
        self.state_var = tk.IntVar(self.newUI)
        self.new_line = tk.IntVar(self.newUI)
        self._count = tk.IntVar(self.newUI)
        self.new_line.set(1)
        self.state_var.set(1)
        self.th_add = tk.LabelFrame(self.newUI,text="Bảng tổng hợp")
        self.th_add.pack(side='left')
        self.th_text = tk.Entry(self.th_add)
        self.th_text.pack()
        self.th_text.insert(0, "A1")
        self.ths_text = tk.Entry(self.th_add)
        self.ths_text.pack()
        self.ths_text.insert(0, self.th_test_sheet)
        self.c_add = tk.LabelFrame(self.newUI,text="Bảng con")
        self.c_add.pack(side='right')
        self.c_text = tk.Entry(self.c_add)
        self.c_text.pack()
        self.c_text.insert(0, "A1:B1")
        self.cs_text = tk.Entry(self.c_add)
        self.cs_text.pack()
        self.cs_text.insert(0, self.c_test_sheet)
        self.ok = tk.Button(self.newUI,text="OK",width=10,command=self.add_data)
        self.ok.pack(side='bottom')
        self._text = tk.Entry(self.newUI)
        self._text.pack(side='top')
        self._text.insert(0, "Vùng "+ str(len(data.datalist)+1))
        self.choosebox = tk.LabelFrame(self.newUI,text="Chế độ")
        self.choosebox.pack()
        self.type_1 = tk.Radiobutton(self.choosebox,text="Tính tổng",variable=self.state_var, value=1)
        self.type_2 = tk.Radiobutton(self.choosebox,text="Đếm",variable=self.state_var, value=2)
        self.type_3 = tk.Radiobutton(self.choosebox,text="Copy",variable=self.state_var, value=3)
        self.setting = tk.LabelFrame(self.newUI,text="Cài đặt")
        self.setting.pack(side='right')
        self.new_line_box = tk.Checkbutton(self.setting,text="Xuống dòng",variable=self.new_line)
        self.new_line_box.pack()
        self.vis_th = tk.Button(self.th_add,text="Chọn",width=10,command=self.visual_sheet_th)
        self.vis_th.pack(side = 'bottom')
        self.vis_c = tk.Button(self.c_add,text="Chọn",width=10,command=self.visual_sheet_c)
        self.vis_c.pack(side = 'bottom')
        self.pp_count = tk.Checkbutton(self.setting,text="Tổng số",variable=self._count)
        self.pp_count.pack()
        self.pp_count.deselect()
        self.pp_count_box = tk.Entry(self.setting)
        self.pp_count_box.pack()
        self.pp_count_box.insert(0,"A1")
        self.pp_count_box['state'] = 'disabled'
        self.pp_count.bind("<Button-1>", self.change_pp_box)
        self.c_text.bind("<Button-1>",self.Update_range)
        self.type_1.pack()
        self.type_2.pack()
        self.type_3.pack()
    def delete_data(self):
        if data.datalist[self.chooselist.curselection()[0]]:
            
            del data.datalist[self.chooselist.curselection()[0]]
            self.chooselist.delete(0,len(data.datalist)) # clear
            for key in data.datalist:
              self.chooselist.insert(data.datalist.index(key)+1 , key[0])
    def add_data(self):
        print(self.c_text.get())
        data.datalist.append([self._text.get(),self.ths_text.get(),self.cs_text.get(),self.th_text.get(),self.c_text.get(),self.state_var.get(),self.new_line.get()])
        print("Che do: "+ str(self.state_var.get()))
        self.chooselist.delete(0,len(data.datalist)) # clear
        for key in data.datalist:
          self.chooselist.insert(data.datalist.index(key)+1 , key[0])
        

        self.newUI.destroy()
    def run_all(self):
        print("Starting..")
        for task in data.datalist:
            state = task[5]
            if state == 2:
                self.counting(data.datalist.index(task),data)
            elif state == 3:
                self.copy(data.datalist.index(task),data)
            elif state == 1:
                self.sum_f(data.datalist.index(task),data)
        messagebox.showinfo("Thành công","Tổng hợp thành công, đang lưu..")
        sleep(1)
        self.reload()
    def run_once(self):

        print(data.datalist)
        try:
            data_ID = self.chooselist.curselection()[0]
        except :
            messagebox.showerror("Lỗi","Bạn chưa chọn task")
        else:
            state = data.datalist[data_ID][5]
            if state == 2:
                self.counting(self.chooselist.curselection()[0],data)
            elif state == 3:
                self.copy(self.chooselist.curselection()[0],data)
            elif state == 1:
                self.sum_f(self.chooselist.curselection()[0],data)
            messagebox.showinfo("Thành công","Tổng hợp thành công, đang lưu..")
            sleep(1)
            self.reload()
    def copy(self,data_ID,data):
         if data.datalist[data_ID]:
            sum_pos =data.datalist[data_ID][3]
            c_pos = data.datalist[data_ID][4]
            sum_sheet  = data.datalist[data_ID][1]
            c_sheet = data.datalist[data_ID][2]
            data = ""
            sumwb =  openpyxl.load_workbook(filename = filepathdata.file_tong_hop)
            
            for file in filepathdata.file_con_path:
                wb =  openpyxl.load_workbook(filename = file)
                sheet_c = wb[c_sheet]
                try:
                       data.datalist[data_ID][3].split(':')[1][0] 
                except:
                    
                    sumwb[sum_sheet][sum_pos].value =  sheet_c[c_pos].value
                    sum_pos = (sum_pos[0] + str(int(sum_pos[1:]) +1)) if self.new_line.get() == 1 else sum_pos
                else:
                    pass

            try:
                sumwb.save(filepathdata.file_tong_hop)
            except :
                messagebox.showerror("Lỗi","Vui lòng đóng file tổng hợp trước khi lưu")
            print("Success")
            
    def reload(self):
        try:
            os.remove(tempFile)
        except:
            pass
        xlapp = win32com.client.DispatchEx("Excel.Application")
        wb = xlapp.workbooks.open(filepathdata.file_tong_hop)
        wb.RefreshAll()
        wb.Save()
        xlapp.Visible = True

    def sum_f(self,data_ID,data):
        if data.datalist[data_ID]:
            sum_pos =data.datalist[data_ID][3]
            c_pos = data.datalist[data_ID][4]
            sum_sheet  = data.datalist[data_ID][1]
            c_sheet = data.datalist[data_ID][2]
            total_of_tt = 0
            sumwb =  openpyxl.load_workbook(filename = filepathdata.file_tong_hop)
            for file in filepathdata.file_con_path:
                
                total = 0
                
                wb =  openpyxl.load_workbook(filename = file)
                sheet_c = wb[c_sheet]
                try:
                       data.datalist[data_ID][3].split(':')[1][0] 
                except:
                    try:
                       data.datalist[data_ID][4].split(':')[1][0]
                    except :
                        if sheet_c[ data.datalist[data_ID][4]].value:
                            total += sheet_c[ data.datalist[data_ID][4]].value
                    else:
                        for range in sheet_c[data.datalist[data_ID][4]]:
                            for cell in range:
                                if cell.value:
                                
                                 print(cell.value)
                                 try:
                                     total += cell.value
                                 except:
                                     print("Cannot sum")
                    
                    
                    total_of_tt += total
                    sumwb[sum_sheet][sum_pos].value = total if self.new_line.get() == 1 else total_of_tt
                    sum_pos = (sum_pos[0] + str(int(sum_pos[1:]) +1)) if self.new_line.get() == 1 else sum_pos
                else:
                    print("Sum position cannot be a range")
                    total = 0
                    pass
                try:
                    sumwb.save(filepathdata.file_tong_hop)
                except :
                    messagebox.showerror("Lỗi","Vui lòng đóng file tổng hợp trước khi lưu")
                print(total)
                self.reload()

    def counting(self,data_ID,data):
        if data.datalist[data_ID]:
            sum_pos =data.datalist[data_ID][3]
            c_pos = data.datalist[data_ID][4]
            sum_sheet  = data.datalist[data_ID][1]
            c_sheet = data.datalist[data_ID][2]
            total_of_tt = 0
            sumwb =  openpyxl.load_workbook(filename = filepathdata.file_tong_hop)
            for file in filepathdata.file_con_path:
                
                total = 0
                
                wb =  openpyxl.load_workbook(filename = file)
                sheet_c = wb[c_sheet]
                try:
                       data.datalist[data_ID][3].split(':')[1][0] 
                except:
                    try:
                       data.datalist[data_ID][4].split(':')[1][0]
                    except :
                        if sheet_c[ data.datalist[data_ID][4]].value:
                            total += 1
                    else:
                        for C_range in sheet_c[data.datalist[data_ID][4]]:
                            for cell in C_range:
                                if cell.value:
                                
                                 print(cell.value)
                                 total += 1
                    
                    
                    total_of_tt += total
                    sumwb[sum_sheet][sum_pos].value = total if self.new_line.get() == 1 else total_of_tt
                    sum_pos = (sum_pos[0] + str(int(sum_pos[1:]) +1)) if self.new_line.get() == 1 else sum_pos
                else:
                    print("Sum position cannot be a range")
                    total = 0
                    pass
                try:
                    sumwb.save(filepathdata.file_tong_hop)
                except :
                 messagebox.showerror("Lỗi","Vui lòng đóng file tổng hợp trước khi lưu")
                print(total)
    def visual_sheet_th(self):
            wb =  openpyxl.load_workbook(filename = filepathdata.file_tong_hop)
            cell_range = wb[self.ths_text.get()]["A1:L20"]
            name = filepathdata.file_tong_hop.split("\\")
            UI = tk.Tk()
            UI.wm_iconbitmap(tempFile)
            UI.title("Bảng "+name[-1])
            print("Range ")
            print(cell_range)
            self.All_Cell = []
            self.Selected_cell = []
            scrollable_frame = tk.LabelFrame(UI,text="Sheet 1")
            scrollable_frame.pack()
            for column in cell_range:
                column_UI = tk.Frame(scrollable_frame)
                column_UI.pack(side='top')
                for cell in column:
                    print("Loading cell: "+cell.coordinate)
                    cell_UI = tk.Entry(column_UI, text=str(cell.value),width=10)
                    cell_UI.delete(0, 'end')
                    cell_UI.insert(0,(str(cell.value) if cell.value != None else ""))
                    cell_UI['state'] = 'disabled'
                    cell_UI['disabledbackground'] = "white"
                    cell_UI.pack(side='left')
                    Cell_data = [cell_UI,cell.coordinate,[cell.column,cell.row]]
                    self.All_Cell.append(Cell_data)
                    cell_UI.bind("<Button-1>",lambda event, arg=cell_UI,pos = self.All_Cell.index(Cell_data):  self.Select_cell(event,arg,pos))
                    


    def visual_sheet_c(self):
            wb =  openpyxl.load_workbook(filename = filepathdata.file_con_path[0])
            cell_range = wb[self.cs_text.get()]["A1:L20"]
            name = filepathdata.file_tong_hop.split("\\")
            UI = tk.Tk()
            UI.wm_iconbitmap(tempFile)
            UI.title("Bảng "+name[-1])
            #container = ttk.Frame(UI)
            #canvas = tk.Canvas(container)
            #scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
            #scrollable_frame = ttk.Frame(canvas)
            #scrollable_frame.bind(
            #"<Configure>",
            #lambda e: canvas.configure(
            #    scrollregion=canvas.bbox("all")
            #    )
            #)
            #scrollbarh = ttk.Scrollbar(container, orient="horizontal", command=canvas.xview)

            #canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            #canvas.configure(yscrollcommand=scrollbar.set)
            #canvas.configure(xscrollcommand=scrollbarh.set)
            print("Range ")
            print(cell_range)
            self.All_Cell = []
            self.Selected_cell = []
            scrollable_frame = tk.LabelFrame(UI,text="Sheet 1")
            scrollable_frame.pack()
            self.holding_mouse = False
            for column in cell_range:
                column_UI = tk.Frame(scrollable_frame)
                column_UI.pack(side='top')
                for cell in column:
                    print("Loading cell: "+cell.coordinate)
                    cell_UI = tk.Entry(column_UI, text=str(cell.value),width=10)
                    cell_UI.delete(0, 'end')
                    cell_UI.insert(0,(str(cell.value) if cell.value != None else ""))
                    cell_UI['state'] = 'disabled'
                    cell_UI['disabledbackground'] = "white"
                    cell_UI.pack(side='left')
                    Cell_data = [cell_UI,cell.coordinate,[cell.column,cell.row]]
                    self.All_Cell.append(Cell_data)
                    cell_UI.bind("<Button-1>",lambda event, arg=cell_UI,pos = self.All_Cell.index(Cell_data),press = True:  self.Select_range(event,arg,pos,press))
                    
            #container.pack( expand=True)
            #canvas.pack(side="left", fill="both", expand=True)
            #scrollbar.pack(side="right", fill="y")
            #scrollbarh.pack(side="bottom", fill="x")
    def Update_range(self,event):
        opened = False
        try:
            self.c_text.get().split(':')[1]
            self.All_Cell
        except :
            pass
        else:
            opened = True
        if opened :
            wb =  openpyxl.load_workbook(filename = filepathdata.file_con_path[0])
            try:        
                Cell_List = wb[self.cs_text.get()][self.c_text.get()]
            except:
                     messagebox.showerror("Lỗi","Kiểu dữ liệu sai")
            else:
             for row in Cell_List:
                    for cell in row:
                        for Cell in self.All_Cell:
                         if Cell[2][1] == cell.row and Cell[2][0] == cell.column :
                             Cell[0]['disabledbackground'] = "#deffe3"
    def Select_range(self,event,arg,pos,press):
             for Cell_data in self.All_Cell:
                 Cell_data[0]['disabledbackground'] = "white"
             if not (arg in self.Selected_cell):
                 self.start_cell_cor = ""
                 self.Selected_cell.clear()
                 self.Selected_cell.append(arg)
                 print(self.Selected_cell)
                 self.c_text.delete(0, 'end')
                 self.c_text.insert(0,self.All_Cell[pos][1])
                 arg['disabledbackground'] = "#deffe3"
                 self.start_cell = self.All_Cell[pos][2]
                 self.start_cell_cor = self.All_Cell[pos][1]
    def Select_cell(self,event,arg,pos):
        if arg in self.Selected_cell:
            self.Selected_cell.remove(arg)
            arg['disabledbackground'] = "white"
            
            for Cell_data in self.All_Cell:
                 Cell_data[0]['disabledbackground'] = "white"
        else:
            for Cell_data in self.All_Cell:
                 Cell_data[0]['disabledbackground'] = "white"
            self.Selected_cell.clear()
            self.Selected_cell.append(arg)
            print(arg.get())
            current_pos = self.All_Cell[pos][2][1]
            for cell in self.Selected_cell:
                cell['disabledbackground'] = "#deffe3"
            if self.new_line.get() == 1 and  len(filepathdata.file_con_path) > 1:
                for x in range(1,len(filepathdata.file_con_path)):
                    print("File so: "+str(x))
                    for Cell_data in self.All_Cell:
                        if Cell_data[2][1]==current_pos + 1 and Cell_data[2][0]  ==self.All_Cell[pos][2][0]:
                            Cell_data[0]['disabledbackground'] = "#dcfaf6"
                            current_pos += 1
                            break
            
        print(self.Selected_cell)
        self.th_text.delete(0, 'end')
        self.th_text.insert(0,self.All_Cell[pos][1])
    def quit(self):
        print("Quit")
        self.active = False
        #self.mn.destroy()
        try:
            os.remove(tempFile)
        except:
            pass
        exit()





            










ImportUI()
print("End")
Main()